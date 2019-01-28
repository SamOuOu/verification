from pymongo import MongoClient
import time
import datetime #透過這個來抓取time
# from bson import ObjectId #透過這個來抓取_id
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import sys
def check():
    excel_filename = input('請輸入要校驗之excel檔名(含副檔名)：\n')
    wb = load_workbook(filename = excel_filename, keep_vba=True)
    sheet = wb.get_sheet_by_name('IBL表格放置區')
    time_after = sheet['B1'].value
    time_before = time_after + datetime.timedelta(days=1)
    station_code = input('請輸入校驗之觀測測站代碼，例如板橋站則輸入466880：\n')
    while True:
        ask = input('是否需要修改雨量及觀測資料來源資料庫夾?(y/n)：\n')
        if ask == 'y':
            rainfolder_name = input('請輸入資料庫中的雨量資料夾名，預設observation_new：\n')
            otherfolder_name = input('請輸入資料庫中的其餘觀測資料夾名，預設observation：\n')
            break
        elif ask == 'n':
            rainfolder_name = 'observation_new'
            otherfolder_name = 'observation'
            break
    #上面設定時間範圍
    uri = "mongodb://heimdall:data3.14info@35.163.6.173" 
    client = MongoClient(uri)
    #上面連線
    db = client["weather"]# 資料庫名
    return (time_after,time_before,station_code,rainfolder_name,otherfolder_name,excel_filename,db)
def get_rain(time_after,,time_before,station_code,rainfolder_name,db):
    rain_collection = db[rainfolder_name]#資料夾名
    #上面找資料庫的資料夾
    timecheck = []
    rain_use = []
    #設定接下來要用的空清單
    rain_target = rain_collection.find({"observation_time" : {"$lte": time_before, "$gte": time_after} , 'station_code' : station_code})
    #上面抓取資料
    print("雨量觀測資料筆數為",rain_collection.count_documents({"observation_time" : {"$lte": time_before, "$gte": time_after} , 'station_code' : station_code}))
    if rain_collection.count_documents({"observation_time" : {"$lte": time_before, "$gte": time_after} , 'station_code' : station_code}) == 0:
        sys.exit("查無雨量資料！")
    #上面看抓到多少資料並檢查是否有資料，若無則跳出
    else:
        for post in rain_target:
            if post['observation_time'] not in timecheck and post['observation_time'].minute == 0:
                timecheck.append(post['observation_time'])
                rain_use.append(post['pcpn']['value'])
                print(post['observation_time'])
            else:
                pass
    rain_use.reverse()
    return(rain_use)
    #以observaton_time檢視並透過分鐘值=0篩選資料，後將其pcpn底下value的值裝入rain_use清單
######################上面處理雨量資料###########################
def get_other(time_before,time_after,station_code,otherfolder_name,db):
    temp_collection = db[otherfolder_name]#資料夾名
    #上面設定抓取資料夾
    timecheck = []
    temp_use = []
    temp_target = temp_collection.find({"observation_time" : {"$lte": time_before, "$gte": time_after} , 'station_code' : station_code})
    #上面抓取資料
    print("溫度觀測資料筆數為",temp_collection.count_documents({"observation_time" : {"$lte": time_before, "$gte": time_after} , 'station_code' : station_code}))
    if temp_collection.count_documents({"observation_time" : {"$lte": time_before, "$gte": time_after} , 'station_code' : station_code}) == 0:
        sys.exit("查無觀測資料！")
    #上面看抓到多少資料並檢查是否有資料，若無則跳出
    else:
        for post in temp_target:
            if post['observation_time'] not in timecheck and post['observation_time'].minute == 0:
                timecheck.append(post['observation_time'])
                temp_use.append(post['temperature']['value'])
            else:
                pass
    temp_use.reverse()
    #以observaton_time檢視及篩選資料並將其temperature底下value的值裝入temp_use清單######################上面處理溫度資料###########################
    humi_collection = db[otherfolder_name]#資料夾名
    #上面設定抓取資料夾
    timecheck = []
    humi_use = []
    humi_target = humi_collection.find({"observation_time" : {"$lte": time_before, "$gte": time_after} , 'station_code' : station_code})
    #上面抓取資料
    print("濕度觀測資料筆數為",humi_collection.count_documents({"observation_time" : {"$lte": time_before, "$gte": time_after} , 'station_code' : station_code}))
    if humi_collection.count_documents({"observation_time" : {"$lte": time_before, "$gte": time_after} , 'station_code' : station_code}) == 0:
        sys.exit("查無觀測資料！")
    #上面看抓到多少資料並檢查是否有資料，若無則跳出
    else:
        for post in humi_target:
            if post['observation_time'] not in timecheck and post['observation_time'].minute == 0:
                timecheck.append(post['observation_time'])
                humi_use.append(post['relative_humidity']['value'])
            else:
                pass
    humi_use.reverse()
    #以observaton_time檢視及篩選資料並將其relative_humidity底下value的值裝入humi_use清單
    ######################上面處理濕度資料###########################
    wind_collection = db[otherfolder_name]#資料夾名
    #上面設定抓取資料夾
    timecheck = []
    wind_use = []
    #上面設定抓取測站
    wind_target = wind_collection.find({"observation_time" : {"$lte": time_before, "$gte": time_after} , 'station_code' : station_code})
    #上面抓取資料
    print("風速觀測資料筆數為",wind_collection.count_documents({"observation_time" : {"$lte": time_before, "$gte": time_after} , 'station_code' : station_code}))
    if wind_collection.count_documents({"observation_time" : {"$lte": time_before, "$gte": time_after} , 'station_code' : station_code}) == 0:
        sys.exit("查無觀測資料！")
    #上面看抓到多少資料並檢查是否有資料，若無則跳出
    else:
        for post in wind_target:
            if post['observation_time'] not in timecheck and post['observation_time'].minute == 0:
                timecheck.append(post['observation_time'])
                wind_use.append(post['wind_speed']['value'])
            else:
                pass
    wind_use.reverse()
    return(temp_use,humi_use,wind_use)
    # 以observaton_time檢視及篩選資料並將其wind_speed底下value的值裝入wind_use清單
    # #######################上面處理風力資料#######################
#########################全部資料##############################
def get_forecast(excel_filename):
    wb = load_workbook(filename = excel_filename, keep_vba=True)
    sheet = wb.get_sheet_by_name('IBL表格放置區')
    #上面找檔案跟資料頁
    time_range = []
    rain = []
    temp = []
    humi = []
    wind = []
    rain_per = []
    for row_cell in sheet['K3':'K26']:
        for cell in row_cell:
            rain.append(cell.value)
    for row_cell in sheet['H3':'H26']:
        for cell in row_cell:
            temp.append(cell.value)
    for row_cell in sheet['I3':'I26']:
        for cell in row_cell:
            humi.append(cell.value)
    for row_cell in sheet['G3':'G26']:
        for cell in row_cell:
            wind.append(cell.value)
    wb = load_workbook(filename = excel_filename, keep_vba=True)
    sheet = wb.get_sheet_by_name('表單輸入區')
    for row_cell in sheet['F6':'AC6']:
        for cell in row_cell:
            time_range.append(cell.value.hour)
    for row_cell in sheet['F12':'AC12']:
        for cell in row_cell:
            rain_per.append(cell.value)
    return(time_range,rain,temp,humi,wind,wb)
    #上面一大段從[指定範圍位置]取得各單項資料
#########################################################
def campare(rain_use,temp_use,humi_use,wind_use,rain,temp,humi,wind):
    rain_diff = []
    temp_diff = []
    humi_diff = []
    wind_diff = []
    rain_abs = []
    temp_abs = []
    humi_abs = []
    wind_abs = []
    ############################
    a = 0
    for i in range(24):
        rain_diff.append(round(rain_use[a] - rain[a],2))
        rain_abs.append(abs(round(rain_use[a] - rain[a],2)))
        a += 1
    ############################
    a = 0
    for i in range(24):
        temp_diff.append(round(temp_use[a] - temp[a],2))
        temp_abs.append(abs(round(temp_use[a] - temp[a],2)))
        a += 1
    ############################
    a = 0
    for i in range(24):
        humi_diff.append(round(humi_use[a] - humi[a],2))
        humi_abs.append(abs(round(humi_use[a] - humi[a],2)))
        a += 1
    ############################
    a = 0
    for i in range(24):
        wind_diff.append(round(wind_use[a] - wind[a],2))
        wind_abs.append(abs(round(wind_use[a] - wind[a],2)))
        a += 1
    return(rain_abs,temp_abs,humi_abs,wind_abs)
################################
def put_in(time_range,rain,temp,humi,wind,rain_use,temp_use,humi_use,wind_use,rain_abs,temp_abs,humi_abs,wind_abs,wb,excel_filename):
    ws2 = wb.create_sheet("校驗結果")
    ali = Alignment(horizontal='center', vertical='center')
    tiktok = range(25)

    ws2['A1'] = '時間'
    ws2['A2'] = '觀測降雨'
    ws2['A3'] = '預報降雨'
    ws2['A4'] = '絕對誤差'
    ws2['A6'] = '觀測溫度'
    ws2['A7'] = '預報溫度'
    ws2['A8'] = '絕對誤差'
    ws2['A10'] = '觀測濕度'
    ws2['A11'] = '預報濕度'
    ws2['A12'] = '絕對誤差'
    ws2['A14'] = '觀測風力'
    ws2['A15'] = '預報風力'
    ws2['A16'] = '絕對誤差'
    for w in range(66,90):  
        for column in ws2.iter_cols():
            a = str(chr(w))
        ws2.column_dimensions[a].width = 5
    a = 0
    for col in range(2, 26):
        _ = ws2.cell(column=col, row=1, value=time_range[a])
        a +=1
    ###################
    a = 0
    for col in range(2, 26):
        _ = ws2.cell(column=col, row=2, value=rain_use[a])
        a +=1
    a = 0
    for col in range(2, 26):
        _ = ws2.cell(column=col, row=3, value=rain[a])
        a +=1
    a = 0
    for col in range(2, 26):
        _ = ws2.cell(column=col, row=4, value=rain_abs[a])
        a +=1
    ####################
    a = 0
    for col in range(2, 26):
        _ = ws2.cell(column=col, row=6, value=temp_use[a])
        a +=1
    a = 0
    for col in range(2, 26):
        _ = ws2.cell(column=col, row=7, value=temp[a])
        a +=1
    a = 0
    for col in range(2, 26):
        _ = ws2.cell(column=col, row=8, value=temp_abs[a])
        a +=1
    ####################
    a = 0
    for col in range(2, 26):
        _ = ws2.cell(column=col, row=10, value=humi_use[a])
        a +=1
    a = 0
    for col in range(2, 26):
        _ = ws2.cell(column=col, row=11, value=humi[a])
        a +=1
    a = 0
    for col in range(2, 26):
        _ = ws2.cell(column=col, row=12, value=humi_abs[a])
        a +=1
    ####################
    a = 0
    for col in range(2, 26):
        _ = ws2.cell(column=col, row=14, value=wind_use[a])
        a +=1
    a = 0
    for col in range(2, 26):
        _ = ws2.cell(column=col, row=15, value=wind[a])
        a +=1
    a = 0
    for col in range(2, 26):
        _ = ws2.cell(column=col, row=16, value=wind_abs[a])
        a +=1
    ####################
    for row in ws2.iter_rows():
        for cell in row:
            cell.alignment = ali

    wb.save(excel_filename)

def main():
    (time_before,time_after,station_code,rainfolder_name,otherfolder_name,excel_filename,db) = check()
    rain_use = get_rain(time_before,time_after,station_code,rainfolder_name,db)
    (temp_use,humi_use,wind_use) = get_other(time_before,time_after,station_code,otherfolder_name,db)
    (time_range,rain,temp,humi,wind,wb) = get_forecast(excel_filename)
    (rain_abs,temp_abs,humi_abs,wind_abs) = campare(rain_use,temp_use,humi_use,wind_use,rain,temp,humi,wind)
    put_in(time_range,rain,temp,humi,wind,rain_use,temp_use,humi_use,wind_use,rain_abs,temp_abs,humi_abs,wind_abs,wb,excel_filename)

if __name__ == "__main__":
    main()